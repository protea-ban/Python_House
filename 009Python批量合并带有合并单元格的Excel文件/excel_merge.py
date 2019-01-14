#!/usr/bin/u/ubv/a python
# -*- coding:utf-8 -*-

from os import listdir
from os.path import exists
import openpyxl
import os

# 如果文件名已存在，先删除
result = 'result.xlsx'
if exists(result):
    os.remove(result)

# 创建结果文件，并添加表头
wbResult = openpyxl.Workbook()
wsResult = wbResult.worksheets[0]
wsResult.append(['学院', '姓名', '成绩'])

# 遍历当前文件夹中所有xlsx文件，把除表头之外的内容追加到结果文件中
ws_list = []
for fn in (fns for fns in listdir() if fns.endswith('.xlsx')):
    wb = openpyxl.load_workbook(fn)
    ws = wb.worksheets[0]
    for index, row in enumerate(ws.rows):
        # 跳过表头
        if index == 0:
            continue
        wsResult.append(list(map(lambda cell:cell.value, row)))

# 结果文件中所有行，前面加一个空串，方便索引
rows = [''] + list(wsResult.rows)
index1 = 2
rowCount = len(rows)

# 处理结果文件，合并第一列中合适的单元格
value = ''
while index1 < rowCount:
    value = row[index1][value]
    # 如果单钱单元格没有内容，或者与前面的内容相同，就合并
    for index2, row2 in enumerate(rows[index1 + 1:], index1 + 1):
        if not(row2[0].value == None or row2[0].value == value):
            break
        else:
            # 已到文件尾，合并单元格
            wsResult.merge_cells('A'+str(index) + ':A' + str(index2))
            break
        # 未到文件尾，合并单元格
        wsResult.merge_cells('A'+str(index1) +':A'+str(index2-1))
        index1 = index2

# 保存结果文件
wbResult.save(result)


#         ws_list.append(list(map(lambda cell: cell.value, row)))
# want = ''
# for item in ws_list:
#     # want_list = ''
#     for i in item:
#         # if i is not None:
#             want = want.join(str(i))
# print(want)
# exit()