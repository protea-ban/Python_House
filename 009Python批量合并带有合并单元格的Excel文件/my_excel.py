#!/usr/bin/u/ubv/a python
# -*- coding:utf-8 -*-

from os.path import exists
import openpyxl
import os

# 如果文件名已存在，先删除
result = '合并自动机5_result.xlsx'
if exists(result):
    os.remove(result)

# 创建结果文件，并添加表头
wbResult = openpyxl.Workbook()
wsResult = wbResult.worksheets[0]
wsResult.append(['result'])

wb = openpyxl.load_workbook('自动机5_数据结构_BSH.xlsx')
ws = wb.worksheets[0]

# for row in ws.rows:
#     for cell in row:
#         print(cell.value, '\n')

# 将每行所有的值都合并到一起并将所有合并好的值放到result_list中
result_list = []
for index, row in enumerate(ws.rows):
    # 跳过表头
    if index == 0:
        continue
    # wsResult.append(list(str("".join(list(map(lambda cell: cell.value, row))))))
    rs_list = list(map(lambda cell: cell.value, row))
    list_str = "".join(rs_list)
    result_list.append(list_str)
# result_list = list(map(lambda item: list("".join(item)), result_list))
result_col = []
# result_col.append()
for item in result_list:
    tmp_list = []
    tmp_list.append(item)
    result_col.append(tmp_list)
# 将每个值写到第一列中
for item in result_col:
    wsResult.append(item)
# zip_result = tuple(zip(*result_list))
# print(result_col)
# exit()
# wsResult.append(rows)

# for i in range(len(result_list)):
#     wsResult["A%d"% (i + 2) ] = str(result_list[i])
    # print(ws["A%d"%(i + 2)])
    # print(result_list[i])



# 保存结果文件
wbResult.save(result)